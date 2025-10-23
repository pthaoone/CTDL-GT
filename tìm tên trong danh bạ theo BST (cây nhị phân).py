from tkinter import Tk, filedialog #thư viện để mở hộp thoại chọn file
from openpyxl import load_workbook #đã install thư viện openpuxl để đọc file excel

class Node:
    def __init__(self, name):
        self.name = name
        self.left = None
        self.right = None

def insert(root, name): #chèn tên
    if root is None:
        return Node(name)
    if name < root.name:
        root.left = insert(root.left, name)
    elif name > root.name:
        root.right = insert(root.right, name)
    return root

def search(root, name):
    if root is None:
        return False
    if name == root.name:
        return True
    elif name < root.name:
        return search(root.left, name)
    else:
        return search(root.right, name)

def read_names_from_excel(filename): #hàm đọc tên từ file excel
    wb = load_workbook(filename) #mở file
    sheet = wb.active 
    names = [] 

    for row in sheet.iter_rows(min_row=2, values_only=True):  # Bỏ qua tiêu đề
        if row[0]:
            names.append(str(row[0]).strip())
    return names

if __name__ == "__main__":
    print("=== TÌM TÊN TRONG DANH BẠ DẠNG CÂY (CHỌN FILE EXCEL) ===")

    root_tk = Tk() #tạo cửa sổ chính của tkinter
    root_tk.withdraw() # Ẩn cửa sổ chính

    # Mở hộp thoại để chọn file
    filename = filedialog.askopenfilename(
        title="Chọn file Excel danh bạ",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )

    if not filename:
        print("Bạn chưa chọn file nào, chương trình kết thúc.")
        exit()

    print(f"File được chọn: {filename}")

    # Đọc danh sách tên
    try:
        names = read_names_from_excel(filename)
        print(f"Đọc được {len(names)} tên từ file.")
    except Exception as e:
        print("Lỗi khi đọc file:", e)
        exit()

    # Tạo cây nhị phân
    root = None
    for name in names:
        root = insert(root, name)

    # Tìm kiếm
    target = input("\nNhập tên cần tìm: ").strip()
    if search(root, target):
        print(f"Tìm thấy '{target}' trong danh bạ.")
    else:
        print(f"Không tìm thấy '{target}' trong danh bạ.")
